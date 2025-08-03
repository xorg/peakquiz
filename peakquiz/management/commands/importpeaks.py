import time
import cloudinary
import cloudinary.uploader
from cloudinary.utils import cloudinary_url
from django.db import IntegrityError
from django.conf import settings
import openpyxl
from requests.adapters import HTTPAdapter
import requests
from django.core.management.base import BaseCommand
from urllib3 import Retry
from peakquiz.models import Peak, Picture


class Command(BaseCommand):
    r2_instance = None

    help = "Imports peaks from database based on input csv file"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):

        super().__init__(stdout, stderr, no_color, force_color)

    def add_arguments(self, parser):
        parser.add_argument("input", type=str)

    def get_picture_url_from_wiki_article(self, url: str):
        """
        Extract main picture from wiki articles like this:
        'https://de.wikipedia.org/wiki/Alphubel'
        Generous sleep to account for annyoing rate limit
        """
        # Wikipedia API query string to get the main image on a page
        # (partial URL will be added to the end)
        query = "http://de.wikipedia.org/w/api.php?action=query&prop=pageimages&format=json&piprop=original&titles="
        partial_url = url.split("/")[-1]
        try:

            data = self.get_data(query + partial_url)
            first_part = data["query"]["pages"]
            # this is a way around not knowing the article id number
            for key, value in first_part.items():
                if value["original"]["source"]:
                    url = value["original"]["source"]
                    title = value.get("title")
                    return url, title
        except Exception as exc:
            print(f"Picture not found for: {partial_url}: {exc}")
            return None

    def get_data(self, url: str) -> dict:
        retries = Retry(total=5, backoff_factor=3, status_forcelist=[429])

        s = requests.Session()
        s.mount("https://", HTTPAdapter(max_retries=retries))
        api_res = s.get(url)
        data = api_res.json()
        return data

    def upload_picture_to_cdn(self, url: str, file_name: str) -> str:
        # authenticate to cdn server
        cloudinary.config(
            cloud_name=settings.CLOUDINARY_CLOUD_NAME,
            api_key=settings.CLOUDINARY_API_KEY,
            api_secret=settings.CLOUDINARY_API_SECRET,
            secure=True,
        )
        # Upload an image
        try:
            upload_result = cloudinary.uploader.upload(url, public_id=file_name)
        except cloudinary.exceptions.BadRequest as e:
            print(f"CDN upload error: {e}")
            return None

        # Optimize delivery by resizing and applying auto-format and auto-quality
        optimize_url, _ = cloudinary_url(file_name, fetch_format="auto", quality="auto")
        return optimize_url, upload_result["asset_id"]

    def get_picture_url_from_wiki_link(self, url: str) -> tuple[str, str, str]:
        """
        Extract picture url from a picture page like this:
        https://de.wikipedia.org/wiki/Datei:Churfirsten_mit_Sommerschnee.JPG

        Returns a tuple like this (original_url, author, title)
        """
        image_filename = url.split(":")[-1]
        api_url = f"https://api.wikimedia.org/core/v1/commons/file/File:{image_filename}"
        data = self.get_data(api_url)
        return data.get("original").get("url"), data.get("latest", {}).get("user", {}).get("name"), data.get("title")

    def handle(self, *args, **options):
        # read xlsx
        wb = openpyxl.load_workbook(options["input"])
        ws = wb["Tabellenblatt1"]
        # This will fail if there is no hyperlink to target
        # 1 - Bild
        # 2 - Gipfel
        # 3 - Höhe
        # 4 - Typ (Haupt / Nebengipfel)
        # 5 - Lage
        # 6 - Gebirge
        # 7 - Dominanz
        # 8 - Schartenhöhe
        for row_num in range(2, len(list(ws.rows))):
            try:
                peak_picture = ws.cell(row=row_num, column=1).hyperlink.target
            except AttributeError:
                peak_picture = None
            peak_name = ws.cell(row=row_num, column=2).value
            peak_url = ws.cell(row=row_num, column=2).hyperlink.target
            height = ws.cell(row=row_num, column=3).value
            peak_type = ws.cell(row=row_num, column=4).value
            region = ws.cell(row=row_num, column=5).value
            mountain_range = ws.cell(row=row_num, column=6).value
            try:
                dominance_distance, dominance_peak = tuple(ws.cell(row=row_num, column=7).value.split("\n"))
            except AttributeError:
                dominance_peak = None
            try:
                prominence_distance, prominence_peak = tuple(ws.cell(row=row_num, column=8).value.split("\n"))
            except AttributeError:
                prominence_distance = None
                if isinstance(ws.cell(row=row_num, column=8).value, float) or isinstance(
                    ws.cell(row=row_num, column=8).value, int
                ):
                    prominence_distance = ws.cell(row=row_num, column=8).value
                prominence_peak = None

            try:
                dominance_distance = float(dominance_distance.replace(",", "."))
            except (ValueError, TypeError, AttributeError):
                dominance_distance = None

            if isinstance(dominance_distance, str):
                dominance_distance = float(dominance_distance.replace(",", "."))

            if isinstance(prominence_distance, str):
                prominence_distance = float(prominence_distance.replace(",", "."))

            if "\n" in region:
                region = region.split("\n")[0]

            peak = Peak(
                name=peak_name,
                elevation=height if type(height) in [int, float] else height.split(" ")[0],
                region=region,
                dominance_distance=dominance_distance,
                dominance_peak=dominance_peak,
                prominence_distance=prominence_distance,
                prominence_peak=prominence_peak,
                peak_type=peak_type,
                mountain_range=mountain_range,
            )
            try:
                peak.save()
            except IntegrityError:
                print(f"{peak_name} already exists, updating")
                peak.id = Peak.objects.get(name=peak_name).id
                peak.save(
                    update_fields=[
                        "elevation",
                        "region",
                        "dominance_distance",
                        "dominance_peak",
                        "prominence_distance",
                        "prominence_peak",
                    ]
                )

            if peak_picture:
                if "wiki/Datei:" in peak_picture:
                    url, author, title = self.get_picture_url_from_wiki_link(peak_picture)

                first_pic = Picture(original_url=url, author=author, title=title, peak=peak)
                print(f"found first picture for {peak_name}: {url}")
                try:
                    first_pic.save()
                except IntegrityError:
                    first_pic.id = Picture.objects.filter(original_url=url).first().id
                    first_pic.save()
                    print(f"Picture {first_pic.original_url} already exists, updating")

                # upload picture to cdn and save again
                cdn_data = self.upload_picture_to_cdn(url, str(first_pic.id))
                if cdn_data:
                    cdn_url, cdn_asset_id = cdn_data
                    first_pic.cdn_url = cdn_url
                    first_pic.cdn_asset_id = cdn_asset_id
                    first_pic.save()

            data = self.get_picture_url_from_wiki_article(peak_url)
            if data:
                second_pic, second_pic_title = data
            if data and second_pic and not second_pic == first_pic.original_url:
                print(f"found second picture for {peak_name}: {second_pic}")
                second_picture = Picture(original_url=second_pic, peak=peak, title=second_pic_title)

                try:
                    second_picture.save()

                except IntegrityError:
                    second_picture.id = Picture.objects.filter(original_url=second_pic).first().id
                    second_picture.save()
                    print(f"Picture {second_picture.original_url} already exists, updating")
                # upload picture to cdn and save again
                cdn_data = self.upload_picture_to_cdn(second_pic, str(second_picture.id))
                if cdn_data:
                    cdn_url, cdn_asset_id = cdn_data
                    second_picture.cdn_url = cdn_url
                    second_picture.cdn_asset_id = cdn_asset_id

                    second_picture.save()
